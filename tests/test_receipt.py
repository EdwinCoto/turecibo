import pytest
import asyncio
from typing import cast
from datetime import date
from io import BytesIO
from openpyxl import load_workbook
from models.receipt import ExtractionData, Receipt, ReceiptSource, ReceiptStatus
from handlers.telegram_handler import _build_excel, _month_name_es, cmd_delete, cmd_excel, cmd_global, cmd_recibo, cmd_sync, handle_text_message
from telegram import Update
from telegram.ext import ContextTypes
from services import vision
from handlers.receipt_handler import (
    _format_missing_emission_date_message,
    _process_receipt,
    _format_duplicate_message,
    _format_invalid_ruc_message,
    _format_missing_dni_message,
    _format_success_message,
)


def make_source(**kwargs) -> ReceiptSource:
    defaults = dict(
        telegram_user_id=1,
        telegram_chat_id=1,
        telegram_message_id=1,
        telegram_file_id="abc",
        telegram_file_unique_id=None,
    )
    return ReceiptSource(**{**defaults, **kwargs})


def test_receipt_defaults():
    r = Receipt(source=make_source())
    assert r.status == ReceiptStatus.PENDING
    assert r.id
    assert r.extraction.data is None


def test_receipt_json_round_trip():
    r = Receipt(source=make_source())
    data = r.to_json_dict()
    assert data["status"] == "pending"
    assert data["extraction"]["status"] == "pending"


def test_extraction_data_currency_default():
    d = ExtractionData(total_amount=59.0)
    assert d.currency == "PEN"


def test_extraction_data_accepts_emission_date():
    d = ExtractionData(emission_date=date(2026, 6, 8))
    assert d.emission_date == date(2026, 6, 8)


def test_extraction_data_accepts_electronic_receipt_number():
    d = ExtractionData(electronic_receipt_number="BPE1-000237")
    assert d.electronic_receipt_number == "BPE1-000237"


# ──────────────────────────────────────────
# DNI validator (sync format checks only)
# ──────────────────────────────────────────

from services.dni_validator import is_valid_format, validate_dni
from services.electronic_receipt_validator import is_valid_format as is_valid_electronic_receipt_format
from services.electronic_receipt_validator import validate_electronic_receipt_number
from services.ruc_validator import is_valid_format as is_valid_ruc_format
from services.ruc_validator import validate_ruc
from storage import azure_backend, google_drive_backend


@pytest.mark.parametrize("dni,expected", [
    ("12345678", True),
    ("00000000", True),
    ("1234567",  False),  # 7 digits
    ("123456789", False), # 9 digits
    ("1234567a",  False), # non-numeric
    ("",          False),
])
def test_dni_format(dni, expected):
    assert is_valid_format(dni) == expected


@pytest.mark.parametrize("dni,expected", [
    ("12345678", True),
    ("00000000", True),
    ("1234567", False),
    ("abc", False),
])
def test_validate_dni_uses_local_format_only(dni, expected):
    assert asyncio.run(validate_dni(dni)) is expected


@pytest.mark.parametrize("ruc,expected", [
    ("20613724851", True),
    ("20605899286", True),
    ("20510885229", True),
    ("20425476115", True),
    ("2061372485", False),
    ("206137248512", False),
    ("2061372485A", False),
    ("20605899287", False),
    ("00000000000", False),
    ("", False),
])
def test_ruc_format(ruc, expected):
    assert is_valid_ruc_format(ruc) == expected


@pytest.mark.parametrize("ruc,expected", [
    ("20613724851", True),
    ("20605899286", True),
    ("20510885229", True),
    ("20425476115", True),
    ("20605899287", False),
    ("2061372485", False),
    ("00000000000", False),
    ("abc", False),
])
def test_validate_ruc_uses_local_format_only(ruc, expected):
    assert asyncio.run(validate_ruc(ruc)) is expected


def test_azure_receipt_date_str_prefers_extraction_emission_date():
    r = Receipt(source=make_source())
    r.receipt_date = None
    r.extraction.data = ExtractionData(emission_date=date(2026, 6, 8))

    assert azure_backend._receipt_date_str(r) == "2026-06-08"


@pytest.mark.parametrize("receipt_number,expected", [
    ("B130-00274475", True),
    ("BPE1-000237", True),
    ("B001-1", True),
    ("F130-00274475", False),
    ("B13-00274475", False),
    ("B130-00000000", False),
    ("B13000274475", False),
    ("", False),
])
def test_electronic_receipt_number_format(receipt_number, expected):
    assert is_valid_electronic_receipt_format(receipt_number) == expected


@pytest.mark.parametrize("receipt_number,expected", [
    ("B130-00274475", True),
    ("BPE1-000237", True),
    ("B130-00000000", False),
    ("F130-00274475", False),
])
def test_validate_electronic_receipt_number(receipt_number, expected):
    assert asyncio.run(validate_electronic_receipt_number(receipt_number)) is expected


# ──────────────────────────────────────────
# Storage helpers
# ──────────────────────────────────────────

import json
import tempfile
from pathlib import Path
from unittest.mock import patch
from unittest.mock import AsyncMock
from types import SimpleNamespace

from storage import receipt_store, local_backend


def test_save_and_load_receipt(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        path = receipt_store.save_receipt(r)
        assert path.exists()
        loaded = json.loads(path.read_text())
        assert loaded["id"] == r.id


def test_get_receipts_by_month_empty(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        result = receipt_store.get_receipts_by_month("2024-01")
        assert result == []


def test_get_receipts_by_month(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        receipt_store.save_receipt(r)
        month = r.created_at.strftime("%Y-%m")
        results = receipt_store.get_receipts_by_month(month)
        assert len(results) == 1
        assert results[0]["id"] == r.id


def test_get_receipts_by_year(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        receipt_store.save_receipt(r)
        year = r.created_at.strftime("%Y")
        results = receipt_store.get_receipts_by_year(year)
        assert len(results) == 1
        assert results[0]["id"] == r.id


def test_get_receipt_by_id(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        receipt_store.save_receipt(r)
        found = receipt_store.get_receipt_by_id(r.id[:8])
        assert found is not None
        assert found["id"] == r.id


def test_get_receipt_by_telegram_file_id(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source(telegram_file_id="file-123"))
        receipt_store.save_receipt(r)
        found = receipt_store.get_receipt_by_telegram_file_id("file-123")
        assert found is not None
        assert found["source"]["telegram_file_id"] == "file-123"


def test_get_receipt_by_telegram_photo_identity_prefers_unique_id(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source(telegram_file_id="file-123", telegram_file_unique_id="unique-abc"))
        receipt_store.save_receipt(r)
        found = receipt_store.get_receipt_by_telegram_photo_identity("unique-abc", "file-999")
        assert found is not None
        assert found["source"]["telegram_file_unique_id"] == "unique-abc"


def test_get_receipt_by_photo_hash(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        r.photo = __import__("models.receipt", fromlist=["ReceiptPhoto"]).ReceiptPhoto(
            local_path=str(tmp_path / "dummy.jpg"),
            size_bytes=11,
            content_hash="abc123",
        )
        receipt_store.save_receipt(r)
        found = receipt_store.get_receipt_by_photo_hash("abc123")
        assert found is not None
        assert found["photo"]["content_hash"] == "abc123"


def test_build_receipt_fingerprint_is_stable(tmp_path):
    receipt = {
        "created_at": "2026-06-08T00:00:00Z",
        "receipt_date": "2026-06-08",
        "extraction": {
            "data": {
                "restaurant_name": "Pardos Chicken",
                "ruc": "20425476115",
                "total_amount": 108.40,
                "currency": "PEN",
                "dni": "72804567",
            }
        },
    }
    fingerprint_1 = receipt_store.build_receipt_fingerprint(receipt)
    fingerprint_2 = receipt_store.build_receipt_fingerprint(receipt)
    assert fingerprint_1 == fingerprint_2


def test_get_receipt_by_fingerprint(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        r.receipt_date = date(2026, 6, 8)
        r.receipt_fingerprint = "fp-123"
        receipt_store.save_receipt(r)
        found = receipt_store.get_receipt_by_fingerprint("fp-123")
        assert found is not None
        assert found["receipt_fingerprint"] == "fp-123"


def test_get_receipt_by_fingerprint_matches_recomputed_when_stored_is_legacy(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        r.receipt_date = date(2026, 5, 3)
        r.extraction.data = ExtractionData(
            restaurant_name="KG BAR S.A.C.",
            ruc="20613724851",
            total_amount=106.0,
            dni="72804567",
            dni_valid=True,
        )
        r.receipt_fingerprint = "legacy-fingerprint"
        receipt_store.save_receipt(r)

        expected = receipt_store.build_receipt_fingerprint(r.to_json_dict())
        found = receipt_store.get_receipt_by_fingerprint(expected)

        assert found is not None
        assert found["id"] == r.id


def test_parse_gdrive_uri_accepts_both_slash_formats():
    assert google_drive_backend.parse_gdrive_uri("gdrive://photos/2026-06-08/file-123") == (
        "photos",
        "2026-06-08",
        "file-123",
    )
    assert google_drive_backend.parse_gdrive_uri("gdrive:/photos/2026-06-08/file-123") == (
        "photos",
        "2026-06-08",
        "file-123",
    )


def test_move_or_upload_photo_keeps_existing_gdrive_uri():
    uri = "gdrive://photos/2026-06-08/file-123"
    result = google_drive_backend._move_or_upload_photo("2026-06-09", "receipt-1", uri)
    assert result == uri


def test_save_receipt_moves_to_emission_date_directory(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        first_path = receipt_store.save_receipt(r)
        photo_path = receipt_store.save_photo(r.id, r.created_at.strftime("%Y-%m-%d"), b"photo-bytes")
        r.photo = __import__("models.receipt", fromlist=["ReceiptPhoto"]).ReceiptPhoto(
            local_path=str(photo_path),
            size_bytes=len(b"photo-bytes"),
        )
        r.receipt_date = date(2026, 6, 8)

        second_path = receipt_store.save_receipt(r)

        assert first_path.parent.name == r.created_at.strftime("%Y-%m-%d")
        assert second_path.parent.name == "2026-06-08"
        assert second_path.exists()
        assert r.photo is not None
        assert r.photo.local_path.endswith("2026-06-08/%s.jpg" % r.id)
        assert r.photo.content_hash is None


def test_save_receipt_routes_to_google_drive_backend(monkeypatch):
    r = Receipt(source=make_source())
    monkeypatch.setenv("STORAGE_BACKEND", "google_drive")

    called = {"count": 0}

    def _fake_save_receipt(_receipt):
        called["count"] += 1
        return Path("gdrive://receipts/2026-06-08/file-123")

    monkeypatch.setattr(receipt_store.google_drive_backend, "save_receipt", _fake_save_receipt)

    result = receipt_store.save_receipt(r)

    assert called["count"] == 1
    assert str(result).startswith("gdrive:")


def test_get_photo_bytes_routes_gdrive_uris(monkeypatch):
    monkeypatch.setattr(receipt_store.google_drive_backend, "get_photo_bytes", lambda _path: b"photo-bytes")

    assert receipt_store.get_photo_bytes("gdrive://photos/2026-06-08/file-123") == b"photo-bytes"
    assert receipt_store.get_photo_bytes("gdrive:/photos/2026-06-08/file-123") == b"photo-bytes"


def test_delete_receipt_by_id_removes_json_and_photo(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        r = Receipt(source=make_source())
        photo_path = receipt_store.save_photo(r.id, r.created_at.strftime("%Y-%m-%d"), b"photo-bytes")
        r.photo = __import__("models.receipt", fromlist=["ReceiptPhoto"]).ReceiptPhoto(
            local_path=str(photo_path),
            size_bytes=len(b"photo-bytes"),
        )
        json_path = receipt_store.save_receipt(r)

        assert json_path.exists()
        assert Path(r.photo.local_path).exists()

        deleted = receipt_store.delete_receipt_by_id(r.id[:8])

        assert deleted is True
        assert not json_path.exists()
        assert not Path(r.photo.local_path).exists()


def test_delete_receipt_by_id_returns_false_when_not_found(tmp_path):
    with patch.object(local_backend, "BASE_PATH", tmp_path):
        assert receipt_store.delete_receipt_by_id("nope1234") is False


def test_normalize_emission_date_value_from_common_formats():
    assert vision._normalize_emission_date_value("08/06/2026") == date(2026, 6, 8)
    assert vision._normalize_emission_date_value("2026-06-08") == date(2026, 6, 8)


def test_normalize_extraction_payload_maps_fecha_emision_to_emission_date():
    payload = {"FECHA DE EMISION": "08/06/2026"}

    normalized = vision._normalize_extraction_payload(payload)

    assert normalized["emission_date"] == date(2026, 6, 8)


def test_handle_text_message_prompts_for_photo():
    reply_text = AsyncMock()
    update = cast(Update, SimpleNamespace(message=SimpleNamespace(reply_text=reply_text)))
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace())

    asyncio.run(handle_text_message(update, context))

    reply_text.assert_awaited_once_with(
        "📸 Envíame una foto de tu boleta para poder procesarla.",
    )


def test_cmd_global_rejects_invalid_year_format():
    reply_text = AsyncMock()
    update = cast(Update, SimpleNamespace(message=SimpleNamespace(reply_text=reply_text)))
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=["20ab"]))

    asyncio.run(cmd_global(update, context))

    reply_text.assert_awaited_once()
    args, kwargs = reply_text.await_args
    assert "Formato inválido" in args[0]
    assert kwargs.get("parse_mode") == "Markdown"


def test_cmd_excel_rejects_invalid_year_format():
    reply_text = AsyncMock()
    update = cast(Update, SimpleNamespace(message=SimpleNamespace(reply_text=reply_text)))
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=["20ab"]))

    asyncio.run(cmd_excel(update, context))

    reply_text.assert_awaited_once()
    args, kwargs = reply_text.await_args
    assert "Formato inválido" in args[0]
    assert kwargs.get("parse_mode") == "Markdown"


def test_cmd_excel_reports_no_data(monkeypatch):
    reply_text = AsyncMock()
    reply_document = AsyncMock()
    update = cast(
        Update,
        SimpleNamespace(message=SimpleNamespace(reply_text=reply_text, reply_document=reply_document)),
    )
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=["2026"]))

    monkeypatch.setattr("handlers.telegram_handler.get_receipts_by_year", lambda _year: [])

    asyncio.run(cmd_excel(update, context))

    reply_text.assert_awaited_once()
    args, _ = reply_text.await_args
    assert "No hay recibos registrados" in args[0]
    reply_document.assert_not_called()


def test_build_excel_skips_unprocessed_or_empty_rows():
    monthly_data = [
        (
            6,
            [
                {
                    "id": "processed-1",
                    "status": "processed",
                    "receipt_date": "2026-06-16",
                    "created_at": "2026-06-17T12:00:00+00:00",
                    "extraction": {
                        "data": {
                            "ruc": "20613724851",
                            "restaurant_name": "RESTAURANTE UNO S.A.C.",
                            "electronic_receipt_number": "B130-00274475",
                            "total_amount": 59.0,
                            "dni": "12345678",
                        }
                    },
                },
                {
                    "id": "pending-1",
                    "status": "pending",
                    "created_at": "2026-06-17T12:05:00+00:00",
                    "extraction": {"data": None},
                },
            ],
        )
    ]

    excel_file = _build_excel(2026, monthly_data)
    workbook = load_workbook(BytesIO(excel_file.getvalue()))
    sheet = workbook["Junio"]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "20613724851"
    assert rows[0][2] == "B130-00274475"
    assert rows[0][3] == "2026-06-16"


def test_cmd_delete_requires_receipt_id_argument():
    reply_text = AsyncMock()
    update = cast(Update, SimpleNamespace(message=SimpleNamespace(reply_text=reply_text)))
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=[]))

    asyncio.run(cmd_delete(update, context))

    reply_text.assert_awaited_once()
    args, kwargs = reply_text.await_args
    assert "Proporciona el ID" in args[0]
    assert kwargs.get("parse_mode") == "Markdown"


def test_cmd_delete_reports_not_found(monkeypatch):
    reply_text = AsyncMock()
    update = cast(Update, SimpleNamespace(message=SimpleNamespace(reply_text=reply_text)))
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=["abc12345"]))

    monkeypatch.setattr("handlers.telegram_handler.delete_receipt_by_id", lambda _rid: False)

    asyncio.run(cmd_delete(update, context))

    reply_text.assert_awaited_once()
    args, kwargs = reply_text.await_args
    assert "no encontrado" in args[0]
    assert kwargs.get("parse_mode") == "Markdown"


def test_cmd_recibo_includes_boleta_number(monkeypatch):
    reply_text = AsyncMock()
    reply_photo = AsyncMock()
    update = cast(
        Update,
        SimpleNamespace(message=SimpleNamespace(reply_text=reply_text, reply_photo=reply_photo)),
    )
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=["abc12345"]))

    monkeypatch.setattr(
        "handlers.telegram_handler.get_receipt_by_id",
        lambda _rid: {
            "id": "abc12345-0000-0000-0000-000000000000",
            "created_at": "2026-06-16T00:00:00+00:00",
            "status": "processed",
            "photo": {},
            "extraction": {
                "data": {
                    "restaurant_name": "Demo",
                    "ruc": "20123456789",
                    "electronic_receipt_number": "B130-00274475",
                    "total_amount": 42.5,
                    "dni": "12345678",
                    "dni_valid": True,
                }
            },
        },
    )

    asyncio.run(cmd_recibo(update, context))


def test_cmd_sync_keeps_existing_boleta_without_reextracting(monkeypatch):
    reply_text = AsyncMock()
    update = cast(Update, SimpleNamespace(message=SimpleNamespace(reply_text=reply_text)))
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=["abc12345"]))

    monkeypatch.setattr(
        "handlers.telegram_handler.get_receipt_by_id",
        lambda _rid: {
            "id": "abc12345-0000-0000-0000-000000000000",
            "photo": {"local_path": "data/receipts/2026-06-16/abc12345.jpg"},
            "extraction": {
                "data": {"electronic_receipt_number": "B130-00274475"},
            },
            "source": {
                "telegram_user_id": 1,
                "telegram_chat_id": 1,
                "telegram_message_id": 1,
                "telegram_file_id": "f1",
            },
        },
    )

    extract_mock = AsyncMock()
    monkeypatch.setattr("handlers.telegram_handler.vision.extract_receipt_data", extract_mock)

    asyncio.run(cmd_sync(update, context))

    extract_mock.assert_not_called()
    reply_text.assert_awaited_once()
    args, kwargs = reply_text.await_args
    assert "ya tiene número de boleta" in args[0]
    assert "B130-00274475" in args[0]
    assert kwargs.get("parse_mode") == "Markdown"


def test_cmd_sync_updates_boleta_when_missing(monkeypatch):
    reply_text = AsyncMock()
    update = cast(Update, SimpleNamespace(message=SimpleNamespace(reply_text=reply_text)))
    context = cast(ContextTypes.DEFAULT_TYPE, SimpleNamespace(args=["abc12345"]))

    monkeypatch.setattr(
        "handlers.telegram_handler.get_receipt_by_id",
        lambda _rid: {
            "id": "abc12345-0000-0000-0000-000000000000",
            "created_at": "2026-06-16T00:00:00+00:00",
            "status": "processed",
            "source": {
                "telegram_user_id": 1,
                "telegram_chat_id": 1,
                "telegram_message_id": 1,
                "telegram_file_id": "f1",
            },
            "photo": {"local_path": "azure://turecibo-receipts/photos/2026-06-16/abc12345.jpg"},
            "extraction": {
                "status": "success",
                "processed_at": "2026-06-16T00:00:00+00:00",
                "data": {
                    "restaurant_name": "Demo",
                    "ruc": "20123456789",
                    "electronic_receipt_number": None,
                    "total_amount": 42.5,
                    "currency": "PEN",
                    "dni": "12345678",
                    "dni_valid": True,
                },
            },
        },
    )
    monkeypatch.setattr("handlers.telegram_handler.get_photo_bytes", lambda _path: b"fake-image")

    extract_mock = AsyncMock(
        return_value=ExtractionData(
            electronic_receipt_number="B130-00274475",
        )
    )
    monkeypatch.setattr("handlers.telegram_handler.vision.extract_receipt_data", extract_mock)

    validate_mock = AsyncMock(return_value=True)
    monkeypatch.setattr(
        "handlers.telegram_handler.electronic_receipt_validator.validate_electronic_receipt_number",
        validate_mock,
    )

    saved = {}

    def _capture_saved(receipt: Receipt):
        saved["receipt"] = receipt

    monkeypatch.setattr("handlers.telegram_handler.save_receipt", _capture_saved)

    asyncio.run(cmd_sync(update, context))

    extract_mock.assert_awaited_once()
    validate_mock.assert_awaited_once_with("B130-00274475")
    assert "receipt" in saved
    assert saved["receipt"].extraction.data is not None
    assert saved["receipt"].extraction.data.electronic_receipt_number == "B130-00274475"

    reply_text.assert_awaited_once()
    args, kwargs = reply_text.await_args
    assert "Sincronización completada" in args[0]
    assert "B130-00274475" in args[0]
    assert kwargs.get("parse_mode") == "Markdown"


def test_format_success_message_includes_receipt_date():
    receipt = Receipt(source=make_source())
    receipt.receipt_date = date(2026, 6, 8)
    extraction = ExtractionData(
        restaurant_name="Demo",
        ruc="20123456789",
        total_amount=10.0,
        dni="12345678",
        dni_valid=True,
    )

    message = _format_success_message(receipt, extraction)

    assert "📅 Fecha: 2026-06-08" in message


def test_format_success_message_includes_electronic_receipt_number():
    receipt = Receipt(source=make_source())
    extraction = ExtractionData(
        restaurant_name="Demo",
        ruc="20123456789",
        electronic_receipt_number="B130-00274475",
        total_amount=10.0,
        dni="12345678",
        dni_valid=True,
    )

    message = _format_success_message(receipt, extraction)

    assert "🧾 Boleta: `B130-00274475`" in message


def test_format_duplicate_message_includes_storage_notice():
    message = _format_duplicate_message({
        "id": "8abaac29-e5a5-42ca-8a6f-375ad5fd6156",
        "receipt_date": "2026-05-03",
    })

    assert "ya está almacenado" in message
    assert "📅 Fecha: 2026-05-03" in message


def test_format_missing_dni_message_mentions_not_saved():
    message = _format_missing_dni_message("8abaac29-e5a5-42ca-8a6f-375ad5fd6156")

    assert "No encontré el DNI" in message
    assert "No lo guardaré" in message


def test_format_invalid_ruc_message_mentions_not_saved():
    message = _format_invalid_ruc_message("8abaac29-e5a5-42ca-8a6f-375ad5fd6156")

    assert "RUC" in message
    assert "No lo guardaré" in message


def test_format_missing_emission_date_message_mentions_not_saved():
    message = _format_missing_emission_date_message("8abaac29-e5a5-42ca-8a6f-375ad5fd6156")

    assert "fecha de emisión" in message
    assert "No lo guardaré" in message
    assert "Vuelve a tomar la foto" in message


def test_process_receipt_skips_save_when_emission_date_missing(monkeypatch):
    receipt = Receipt(source=make_source())
    photo = cast(object, SimpleNamespace(file_id="photo-file-1", file_unique_id="u1", file_size=100))

    download_file = SimpleNamespace(download_as_bytearray=AsyncMock(return_value=bytearray(b"fake-image")))
    bot = SimpleNamespace(get_file=AsyncMock(return_value=download_file))

    send_message_mock = AsyncMock()
    save_counts = {"photo": 0, "receipt": 0}

    monkeypatch.setattr("handlers.receipt_handler.get_bot", lambda: bot)
    monkeypatch.setattr(
        "handlers.receipt_handler.vision.extract_receipt_data",
        AsyncMock(
            return_value=ExtractionData(
                ruc="20613724851",
                dni="12345678",
                emission_date=None,
            )
        ),
    )
    monkeypatch.setattr("handlers.receipt_handler.send_message", send_message_mock)
    monkeypatch.setattr(
        "handlers.receipt_handler.save_photo",
        lambda *_args, **_kwargs: save_counts.__setitem__("photo", save_counts["photo"] + 1),
    )
    monkeypatch.setattr(
        "handlers.receipt_handler.save_receipt",
        lambda *_args, **_kwargs: save_counts.__setitem__("receipt", save_counts["receipt"] + 1),
    )

    asyncio.run(_process_receipt(receipt, cast(object, photo), chat_id=1))

    send_message_mock.assert_awaited_once()
    msg_args = send_message_mock.await_args.args
    assert msg_args[0] == 1
    assert "No pude extraer la fecha de emisión" in msg_args[1]

    assert save_counts["photo"] == 0
    assert save_counts["receipt"] == 0


def test_month_name_es_uses_spanish_labels():
    assert _month_name_es(5) == "Mayo"
    assert _month_name_es(12) == "Diciembre"


def test_month_name_es_raises_for_invalid_month():
    with pytest.raises(ValueError):
        _month_name_es(0)

    with pytest.raises(ValueError):
        _month_name_es(13)
